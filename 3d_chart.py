import requests
import json
from openpyxl import Workbook
from openpyxl.chart import LineChart3D, Reference, Series
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.colors import ColorChoice


rates = []

for year in range(2020, 2024):
    for month in range(1, 13):
        r = requests.get(f'https://api.privatbank.ua/p24api/exchange_rates?json&date=01.{month:02}.{year}')
        if r.status_code == 200:
            response = json.loads(r.text)
            rate_usd = [i for i in response['exchangeRate'] if i['currency'] == 'USD']
            rate_euro = [i for i in response['exchangeRate'] if i['currency'] == 'EUR']
            rate_gbp = [i for i in response['exchangeRate'] if i['currency'] == 'GBP']

            date_str = f'01.{month:02}.{year}'
            usd_rate = rate_usd[0]['saleRate'] if rate_usd else None
            euro_rate = rate_euro[0]['saleRate'] if rate_euro else None
            gbp_rate = rate_gbp[0]['saleRate'] if rate_gbp else None


            rates.append((date_str, usd_rate, euro_rate, gbp_rate))


wb = Workbook()
ws = wb.active
ws.title = "Currency Rates"


ws['A1'] = 'Date'
ws['B1'] = 'USD'
ws['C1'] = 'EUR'
ws['D1'] = 'GBP'


for idx, rate in enumerate(rates, start=2):
    ws[f'A{idx}'] = rate[0]
    ws[f'B{idx}'] = rate[1]
    ws[f'C{idx}'] = rate[2]
    ws[f'D{idx}'] = rate[3]


c1 = LineChart3D()
c1.title = "3D Currency Exchange Rates"
c1.style = 13
c1.y_axis.title = 'Exchange Rate'
c1.x_axis.title = 'Date'


usd_color = 'FF5733'  # Red
eur_color = '33FF57'  # Green
gbp_color = '3357FF'  # Blue


data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=len(rates) + 1)
cats = Reference(ws, min_col=1, min_row=2, max_row=len(rates) + 1)


usd_series = Series(data, title="USD")
usd_series.graphical_properties = GraphicalProperties(
    ln=LineProperties(solidFill=usd_color)
)
c1.series.append(usd_series)


eur_series = Series(data, title="EUR")
eur_series.graphical_properties = GraphicalProperties(
    ln=LineProperties(solidFill=eur_color)
)
c1.series.append(eur_series)


gbp_series = Series(data, title="GBP")
gbp_series.graphical_properties = GraphicalProperties(
    ln=LineProperties(solidFill=gbp_color)
)
c1.series.append(gbp_series)


c1.set_categories(cats)


ws.add_chart(c1, "F2")


wb.save("my_rates_3D.xlsx")

