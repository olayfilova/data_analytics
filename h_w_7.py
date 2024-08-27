import json
from openpyxl import Workbook
import requests
from openpyxl.chart import Reference, LineChart
from openpyxl.chart.axis import DateAxis


rates = []
for year in range(2020, 2024):
    for month in range(1, 13):
        r = requests.get(f'https://api.privatbank.ua/p24api/exchange_rates?json&date=01.{month:02}.{year}')
        if r.status_code == 200:
            response = json.loads(r.text)
            rate_usd = [i for i in response['exchangeRate'] if i['currency'] == 'USD']
            if rate_usd:
                print(year, month, rate_usd[0]['saleRate'])
                rates.append((f'01.{month:02}.{year}', rate_usd[0]['saleRate']))

print(rates)

wb = Workbook()
ws = wb.active
ws['A1'] = 'Date'
ws['B1'] = 'Rate'

for idx, rate in enumerate(rates, start=2):
    ws[f'A{idx}'] = rate[0]
    ws[f'B{idx}'] = rate[1]

data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(rates) + 1)
c1 = LineChart()
c1.title = "USD RATE"
c1.style = 13
c1.y_axis.title = 'Rate'
c1.x_axis.title = 'Date'
c1.y_axis.crossAx = 500
c1.x_axis = DateAxis(crossAx=100)
c1.x_axis.number_format = 'dd.mm.yyyy'
c1.x_axis.majorTimeUnit = "months"

c1.add_data(data, titles_from_data=True)
dates = Reference(ws, min_col=1, min_row=2, max_row=len(rates) + 1)
c1.set_categories(dates)

#smooth to False
for series in c1.series:
    series.smooth = False

ws.add_chart(c1, "D1")

wb.save("my_rate.xlsx")



